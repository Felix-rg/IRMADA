from datetime import datetime
from fastapi.staticfiles import StaticFiles
from fastapi import FastAPI, Request, Form
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse, HTMLResponse
from openpyxl import load_workbook
from fastapi.responses import RedirectResponse
from fastapi.responses import RedirectResponse
from fastapi import Request, Form
from starlette.middleware.sessions import SessionMiddleware


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi.responses import FileResponse
from openpyxl.utils import get_column_letter

import sqlite3

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="irmada-secret")
app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")

def db():
    return sqlite3.connect("database.db")


@app.get("/")
def home(request: Request):

    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)

    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/fitrah")
def form_fitrah(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)
    
    return templates.TemplateResponse("fitrah.html", {"request": request})


@app.post("/fitrah")
def simpan_fitrah(
    tanggal:str=Form(...),
    jam = datetime.now().strftime("%H:%M"),
    kategori:str=Form(...),
    nama:str=Form(...),
    alamat:str=Form(...),
    rt:str=Form(...),
    rw:str=Form(...),
    jiwa:int=Form(...),
    bungkus:int=Form(...)
):

    conn = db()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO fitrah
        (tanggal,jam,nama,alamat,rt,rw,kategori,jiwa,bungkus)
        VALUES (?,?,?,?,?,?,?,?,?)
        """,
        (tanggal,jam,nama,alamat,rt,rw,kategori,jiwa,bungkus)
        )

    conn.commit()
    conn.close()

    return RedirectResponse("/fitrah?success=1", status_code=303)


@app.get("/maal")
def form_maal(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)
    
    return templates.TemplateResponse("maal.html", {"request": request})

@app.post("/maal")
def simpan_maal(
    tanggal:str=Form(...),
    kategori:str=Form(...),
    nama:str=Form(...),
    alamat:str=Form(...),
    rt:str=Form(...),
    rw:str=Form(...),
    jenis:str=Form(...),
    nominal:int=Form(...)
):

    jam = datetime.now().strftime("%H:%M")

    conn = db()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO maal
        (tanggal,jam,kategori,nama,alamat,rt,rw,jenis,nominal)
        VALUES (?,?,?,?,?,?,?,?,?)
        """,
        (tanggal,jam,kategori,nama,alamat,rt,rw,jenis,nominal)
    )

    conn.commit()
    conn.close()

    return RedirectResponse("/maal?success=1", status_code=303)

@app.get("/laporan")
def laporan(request: Request):

    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)
    
    conn = db()
    cur = conn.cursor()

    # ambil data tabel
    cur.execute("SELECT * FROM fitrah ORDER BY id DESC")
    data_fitrah = cur.fetchall()

    cur.execute("SELECT * FROM maal ORDER BY id DESC")
    data_maal = cur.fetchall()

    # hitung total
    cur.execute("SELECT SUM(jiwa) FROM fitrah")
    total_jiwa = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal")
    total_maal = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(bungkus) FROM fitrah")
    total_bungkus = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Zakat Maal'")
    total_zakatmaal = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Infaq'")
    total_infaq = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Shodaqoh'")
    total_shodaqoh = cur.fetchone()[0] or 0

    total_semua = total_zakatmaal + total_infaq + total_shodaqoh
    # baru tutup database
    conn.close()

    return templates.TemplateResponse(
        "laporan.html",
        {
            "request": request,
            "fitrah": data_fitrah,
            "maal": data_maal,
            "total_jiwa": total_jiwa,
            "total_bungkus": total_bungkus,

            "total_zakatmaal": total_zakatmaal,
            "total_infaq": total_infaq,
            "total_shodaqoh": total_shodaqoh,
            "total_semua": total_semua
        }
    )

@app.get("/hapus_fitrah/{id}")
def hapus_fitrah(id:int):

    conn = db()
    cur = conn.cursor()

    cur.execute("DELETE FROM fitrah WHERE id=?", (id,))

    conn.commit()
    conn.close()

    return RedirectResponse("/laporan", status_code=303)


@app.get("/hapus_maal/{id}")
def hapus_maal(id:int):

    conn = db()
    cur = conn.cursor()

    cur.execute("DELETE FROM maal WHERE id=?", (id,))

    conn.commit()
    conn.close()

    return RedirectResponse("/laporan", status_code=303)

@app.post("/edit_fitrah/{id}")
def update_fitrah(
    id:int,
    tanggal:str = Form(...),
    nama:str = Form(...),
    alamat:str = Form(...),
    rt:str = Form(...),
    rw:str = Form(...),
    kategori:str = Form(...),
    jiwa:int = Form(...),
    bungkus:int = Form(...)
):

    conn = db()
    cur = conn.cursor()

    cur.execute(
        """
        UPDATE fitrah
        SET tanggal=?, nama=?, alamat=?, rt=?, rw=?, kategori=?, jiwa=?, bungkus=?
        WHERE id=?
        """,
        (tanggal, nama, alamat, rt, rw, kategori, jiwa, bungkus, id)
    )

    conn.commit()
    conn.close()

    return RedirectResponse("/laporan", status_code=303)

@app.post("/edit_maal/{id}")
def update_maal(
    id:int,
    tanggal:str = Form(...),
    nama:str = Form(...),
    alamat:str = Form(...),
    rt:str = Form(...),
    rw:str = Form(...),
    kategori:str = Form(...),
    jenis:str = Form(...),
    nominal:int = Form(...)
):

    conn = db()
    cur = conn.cursor()

    cur.execute(
        """
        UPDATE maal
        SET tanggal=?, nama=?, alamat=?, rt=?, rw=?, kategori=?, jenis=?, nominal=?
        WHERE id=?
        """,
        (tanggal, nama, alamat, rt, rw, kategori, jenis, nominal, id)
    )

    conn.commit()
    conn.close()

    return RedirectResponse("/laporan", status_code=303)


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):

    if username == "irmada" and password == "zakat2026":
        request.session["user"] = username
        return RedirectResponse("/", status_code=303)

    return RedirectResponse("/login?error=1", status_code=303)

@app.get("/login")
def login_page(request: Request):

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT SUM(jiwa) FROM fitrah")
    total_jiwa = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(bungkus) FROM fitrah")
    total_bungkus = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Zakat Maal'")
    total_maal = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Infaq'")
    total_infaq = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Shodaqoh'")
    total_shodaqoh = cur.fetchone()[0] or 0

    conn.close()

    error = request.query_params.get("error")

    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "error": error,
            "total_jiwa": total_jiwa,
            "total_bungkus": total_bungkus,
            "total_maal": total_maal,
            "total_infaq": total_infaq,
            "total_shodaqoh": total_shodaqoh
        }
    )

@app.get("/logout")
def logout(request: Request):

    request.session.clear()

    return RedirectResponse("/login", status_code=303)

@app.get("/export_fitrah")
def export_fitrah():

    conn = db()
    cur = conn.cursor()

    cur.execute("""
    SELECT tanggal,jam,nama,alamat,rt,rw,kategori,jiwa,bungkus
    FROM fitrah
    ORDER BY rt,rw,nama
    """)

    data = cur.fetchall()
    conn.close()

    wb = load_workbook("ZAKAT FITRAH 2026.xlsx")
    ws = wb.active

    row = 7  # mulai dari baris data

    for i, x in enumerate(data, start=1):

        ws[f"A{row}"] = i          # No
        ws[f"B{row}"] = x[0]       # tanggal
        ws[f"C{row}"] = x[1]       # jam
        ws[f"D{row}"] = x[2]       # nama
        ws[f"E{row}"] = x[3]       # alamat
        ws[f"F{row}"] = x[4]       # rt
        ws[f"G{row}"] = x[5]       # rw
        ws[f"H{row}"] = x[6]       # kategori
        ws[f"I{row}"] = x[7]       # jiwa
        ws[f"J{row}"] = x[8]       # beras

        row += 1

    file = "laporan_fitrah.xlsx"
    wb.save(file)

    return FileResponse(
        file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=file
    )

@app.get("/export_maal")
def export_maal():

    conn = db()
    cur = conn.cursor()

    cur.execute("""
    SELECT tanggal,jam,nama,alamat,rt,rw,kategori,jenis,nominal
    FROM maal
    ORDER BY rt,rw,nama
    """)

    data = cur.fetchall()
    conn.close()

    wb = load_workbook("ZAKAT MAAL_INFAQ_SHODAQOH 2026.xlsx")
    ws = wb.active

    row = 7

    for i,x in enumerate(data,start=1):

        ws[f"A{row}"] = i
        ws[f"B{row}"] = x[0]   # tanggal
        ws[f"C{row}"] = x[1]   # jam
        ws[f"D{row}"] = x[2]   # nama
        ws[f"E{row}"] = x[3]   # alamat
        ws[f"F{row}"] = x[4]   # rt
        ws[f"G{row}"] = x[5]   # rw
        ws[f"H{row}"] = x[6]   # kategori
        ws[f"I{row}"] = x[7]   # jenis
        ws[f"J{row}"] = x[8]   # nominal

        row += 1

    file = "laporan_maal.xlsx"
    wb.save(file)

    return FileResponse(
        file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=file
    )

@app.get("/audit_beras")
def audit_beras(request: Request):

    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT SUM(jiwa) FROM fitrah")
    total_jiwa = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(bungkus) FROM fitrah")
    total_beras = cur.fetchone()[0] or 0

    cur.execute("SELECT COUNT(*) FROM fitrah")
    total_transaksi = cur.fetchone()[0]

    conn.close()

    return templates.TemplateResponse(
        "audit_beras.html",
        {
            "request": request,
            "total_jiwa": total_jiwa,
            "total_beras": total_beras,
            "total_transaksi": total_transaksi
        }
    )

@app.get("/audit_uang")
def audit_uang(request: Request):

    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Zakat Maal'")
    zakat = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Infaq'")
    infaq = cur.fetchone()[0] or 0

    cur.execute("SELECT SUM(nominal) FROM maal WHERE jenis='Shodaqoh'")
    shodaqoh = cur.fetchone()[0] or 0

    total = zakat + infaq + shodaqoh

    conn.close()

    return templates.TemplateResponse(
        "audit_uang.html",
        {
            "request": request,
            "zakat": zakat,
            "infaq": infaq,
            "shodaqoh": shodaqoh,
            "total": total
        }
    )


@app.get("/penyaluran")
def penyaluran(request: Request):

    if not request.session.get("user"):
        return RedirectResponse("/login", status_code=303)

    conn = db()
    cur = conn.cursor()

    # total beras masuk dari zakat fitrah
    cur.execute("SELECT SUM(bungkus) FROM fitrah")
    total_beras = cur.fetchone()[0] or 0

    # total uang zakat maal
    cur.execute("SELECT SUM(nominal) FROM maal")
    total_uang = cur.fetchone()[0] or 0

    # total beras yang sudah disalurkan
    cur.execute("SELECT SUM(beras) FROM penyaluran")
    total_disalurkan = cur.fetchone()[0] or 0

    # sisa beras
    sisa_beras = total_beras - total_disalurkan

    # data penerima zakat
    cur.execute("SELECT * FROM penyaluran ORDER BY id DESC")
    data_penyaluran = cur.fetchall()

    conn.close()

    return templates.TemplateResponse(
        "penyaluran.html",
        {
            "request": request,
            "total_beras": total_beras,
            "total_uang": total_uang,
            "total_disalurkan": total_disalurkan,
            "sisa_beras": sisa_beras,
            "penyaluran": data_penyaluran
        }
    )


@app.post("/penyaluran")
def simpan_penyaluran(
    tanggal:str=Form(...),
    nama:str=Form(...),
    alamat:str=Form(...),
    rt:str=Form(...),
    rw:str=Form(...),
    kategori:str=Form(...),
    beras:int=Form(...)
):

    conn = db()
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO penyaluran
        (tanggal,nama,alamat,rt,rw,kategori,beras)
        VALUES (?,?,?,?,?,?,?)
        """,
        (tanggal,nama,alamat,rt,rw,kategori,beras)
    )

    conn.commit()
    conn.close()

    return RedirectResponse("/penyaluran", status_code=303) 